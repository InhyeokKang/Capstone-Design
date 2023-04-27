import blescan
import sys
import bluetooth._bluetooth as bluez
import pymysql
from time import sleep
import threading
import os
import openpyxl
import time

class DB_sending:
    def __init__(self):
        self.url = "210.115.227.108"
        self.id = 'cic'
        self.password = '20180903in'
        self.dbName = 'kindergartenbus'
    def creat_connet(self):
        self.db = pymysql.connect(host=self.url, port=3306, user=self.id, passwd=self.password, db=self.dbName, charset='utf8')
        self.cursor = self.db.cursor()

    def calcualte_distance_rssi(self, txPower, rssi):
        txPower_num =  int(txPower)
        rssi_num = int(rssi)
        if rssi_num ==  0 :
            return -1

        ratio = rssi_num * 1.0 / txPower_num
        if ratio < 1.0 :
            return str(ratio**10)
        else:
            distance = (0.89976) * (ratio**7.7095) + 0.111
            return str(distance)

    def insert_unique_data(self, mac, uuid, major, minor):
        sql = "insert into device_unique_info_tb (macAddress, UUID, major, minor) " \
                "select '"+ mac+"' ,'"+uuid+"' ,'"+major+"' ,'"+minor+"' from dual where not exists" \
                "( select * from device_unique_info_tb where macAddress = '"+mac+"' and UUID = '"+uuid+"')"
        print(sql)
        self.cursor.execute(sql)
        self.db.commit()
        print(self.cursor.lastrowid)

    def insert_valiable_data(self, mac, rssi, txpower, accuracy):
        sql = "INSERT INTO `device_variable_info_tb` (`macaddress`, `rssi`, `txpower`, `accuracy`, `time`) VALUES ('"+ mac +"', '"+ rssi +"', '"+ txpower +"', '"+ accuracy +"', CURRENT_TIMESTAMP);"
        print(sql)
        self.cursor.execute(sql)
        self.db.commit()
        print(self.cursor.lastrowid)

    def run_sensor_thread(self):
        os.system("sudo python3 /home/pi/sensorDataToDB.py")


dev_id = 0
conn = DB_sending()

try:
    sock = bluez.hci_open_dev(dev_id)
    print("ble thread started")

except:
    print("error accessing bluetooth device...")
    sys.exit(1)

blescan.hci_le_set_scan_parameters(sock)
blescan.hci_enable_le_scan(sock)


start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

wb = openpyxl.Workbook()
sheet1 = wb['Sheet']
sheet1.title = 'Collected data'
sheet1.cell(row=1, column=1).value = 'MAC'
sheet1.cell(row=1, column=2).value = 'Major'
sheet1.cell(row=1, column=3).value = 'Minor'
sheet1.cell(row=1, column=4).value = 'RSSI'
sheet1.cell(row=1, column=5).value = 'TX_power'
sheet1.cell(row=1, column=6).value = 'Accuracy'
sheet1.cell(row=1, column=7).value = 'yyyy-mm-dd h:m:s)'
wb.save(start_time + '.xlsx')

SAVE_FILE_INTERVAL = 5  # 데이터 파일 저장 주기 (초)

row_count = 1  # 엑셀 파일에 쓴 행 수

time_check = time.time()

while True:
    returnedList = blescan.parse_events(sock, 10)
    for beacon in returnedList:
        beacon_split = beacon.split(',')
        if beacon_split[3] in ["31142"]:
            sheet1.append([beacon_split[0], beacon_split[2], beacon_split[3], beacon_split[5], beacon_split[4], conn.calcualte_distance_rssi(beacon_split[4],beacon_split[5]), time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))])
            row_count += 1
    
    # 일정 시간이 지나면 데이터를 파일에 저장
    if time.time() - time_check > SAVE_FILE_INTERVAL:
        wb.save(start_time + '.xlsx')
        time_check = time.time()
